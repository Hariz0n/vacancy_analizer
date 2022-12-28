from Statistics import Statistic
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Side, Border
import pdfkit
from jinja2 import Environment, FileSystemLoader


class Report:
    """Класс для составления отчетов о вакансиях

    Attributes:
        work_name (str): Название вакансии
        city_perc (dict(str, int)): Словарь с информацией долях вакансий по городам
        salary_city (dict(str, int)): Словарь с информацией о уровнях зарплат по городам
        count_worker (dict(str, int)): Словарь с динамикой количества вакансий по годам для выбранной профессии
        salary_worker (dict(str, int)): Словарь с динамикой уровня зарплат по годам для выбранной профессии
        count (dict(str, int)): Словарь с динамикой количества вакансий по годам
        salary (dict(str, int)): Словарь с динамикой уровня зарплат по годам
    """

    def __init__(self, stats: Statistic):
        """ Инициализирует объект Report

        :param stats: Статистика
        :type stats: Statistic
        """
        # print(stats.salary_mean, stats.count, stats.worker_mean, stats.worker_count, stats.city_perc, sep='\n\n')
        self.work_name = stats.work_name
        self.salary: dict = stats.salary_mean
        self.salary_worker: dict = stats.worker_mean
        self.count: dict = stats.count
        self.count_worker: dict = stats.worker_count
        self.salary_city: dict = stats.city_salary
        self.city_perc: dict = stats.city_perc

    def generate_excel(self):
        """
        Генерирует Excel таблицу с данными из статистики
        :return: None
        :rtype: None
        """
        wb = Workbook()

        # Статистика по годам
        wb_years = wb.active
        wb_years.title = 'Статистика по годам'
        wb_years.append(['Год ', 'Средняя зарплата', f'Средняя зарплата - {self.work_name}',
                         'Количество вакансий', f'Количество вакансий - {self.work_name}'])
        for year in self.salary:
            wb_years.append([year, self.salary[year], self.salary_worker[year],
                             self.count[year], self.count_worker[year]])

        for i, col in enumerate(wb_years.columns, 1):
            te = max([len(''.join(str(j.value).split())) for j in col]) or 2
            wb_years.column_dimensions[get_column_letter(i)].width = te + 1.58

        # Статистика по городам
        wb_cities = wb.create_sheet('Статистика по городам')
        wb_cities.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        print(self.salary_city, self.city_perc)
        for (city, salary), (city2, perc) in zip(self.salary_city.items(), self.city_perc.items()):
            wb_cities.append([city, salary, '', city2, perc])

        for i, col in enumerate(wb_cities.columns, 1):
            te = max([len(''.join(str(j.value).split())) for j in col]) or 2
            wb_cities.column_dimensions[get_column_letter(i)].width = te + 1.58

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            wb_years[col + '1'].font = font_bold
            wb_cities[col + '1'].font = font_bold

        for i, v in enumerate(self.city_perc, 2):
            wb_cities['E' + str(i)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for col in wb_years:
            for cell in col:
                if cell.value != '':
                    cell.border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for col in wb_cities:
            for cell in col:
                if cell.value != '':
                    cell.border = Border(left=thin, bottom=thin, right=thin, top=thin)

        wb.save('report.xlsx')

    def generate_image(self):
        """
        Создает изображение с графиками
        :return: None
        """
        pos = np.arange(len(self.salary.keys()))
        plt.rc('font', size=8)
        width = 0.35
        fig, ((salary, count), (city_s, city_p)) = plt.subplots(2, 2)

        # Уровень зарплат по годам
        salary.bar(pos - width / 2, self.salary.values(), width, label='Средняя з/п')
        salary.bar(pos + width / 2, self.salary_worker.values(), width, label=f'з/п {self.work_name}')
        salary.set_title('Уровень зарплат по годам', fontsize=12)
        salary.set_xticks(pos, self.salary.keys(), rotation=90)
        salary.legend()
        salary.grid(axis='y')

        # Количество вакансий по годам
        count.bar(pos - width / 2, self.count.values(), width, label='Количество вакансий')
        count.bar(pos + width / 2, self.count_worker.values(), width, label=f'Количество вакансий\n{self.work_name}')
        count.set_title('Количество вакансий по годам', fontsize=12)
        count.set_xticks(pos, self.count.keys(), rotation=90)
        count.locator_params(axis='y', nbins=5)
        count.legend()
        count.grid(axis='y')

        # Уровень зарплат по городам
        y_pos = np.arange(len(self.salary_city.keys()))
        labels = ['\n'.join(i.split()) for i in ['-\n'.join(e.split('-')) for e in self.salary_city.keys()]]
        city_s.barh(y_pos, self.salary_city.values(), align='center')
        city_s.set_yticks(y_pos, labels=labels, fontsize=6, wrap=True)
        city_s.invert_yaxis()
        city_s.locator_params(axis='x', nbins=4)
        city_s.set_title('Уровень зарплат по городам', fontsize=12)
        city_s.grid(axis="x")

        # Доля вакансий по городам
        others = 1 - sum(self.city_perc.values())
        pie_data = self.city_perc.values() if others == 0 else [others] + list(self.city_perc.values())
        labels = self.city_perc.keys() if others == 0 else ['Другие'] + list(self.city_perc.keys())
        city_p.pie(pie_data, labels=labels, textprops={'fontsize': 6})
        city_p.set_title('Доля вакансий по городам', fontsize=12)

        fig.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def generate_pdf(self):
        """
        Создает pdf отчет с таблицами и графиком
        :return: None
        """
        config = pdfkit.configuration(wkhtmltopdf=r'F:\Programs\wkhtmltopdf\bin\wkhtmltopdf.exe')
        env = Environment(loader=FileSystemLoader('.'))
        table_general = f'<tr>' \
                        f'<th>Год</th>' \
                        f'<th>Средняя зарплата</th>' \
                        f'<th>Средняя зарплата - {self.work_name}</th>' \
                        f'<th>Количество вакансий</th>' \
                        f'<th>Количество вакансий - {self.work_name}</th>' \
                        f'</tr>\n'
        table_salary = f'<tr>' \
                       f'<th>Город</th>' \
                       f'<th>Уровень зарплат</th>' \
                       f'</tr>\n'
        table_perc = f'<tr>' \
                     f'<th>Город</th>' \
                     f'<th>Доля вакансий</th>' \
                     f'</tr>\n'
        for year in self.salary:
            table_general += f'<tr>' \
                             f'<td>{year}</th>' \
                             f'<td>{self.salary[year]}</th>' \
                             f'<td>{self.salary_worker[year]}</th>' \
                             f'<td>{self.count[year]}</th>' \
                             f'<td>{self.count_worker[year]}</th>' \
                             f'</tr>\n'
        for city in self.salary_city:
            table_salary += f'<tr>' \
                             f'<td>{city}</th>' \
                             f'<td>{self.salary_city[city]}</th>' \
                             f'</tr>\n'
        for city in self.city_perc:
            table_perc += f'<tr>' \
                         f'<td>{city}</th>' \
                         f'<td>{f"{self.city_perc[city]*100:.2f}"}%</th>' \
                         f'</tr>\n'
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(
            {'work_name': self.work_name,
             'image_file': 'graph.png',
             'table_general': table_general,
             'table_salary': table_salary,
             'table_perc': table_perc})
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': True})
